"""分析结果"""
import os
import sys
import mne
import json
import numpy as np
import pandas as pd
import matplotlib as mpl
from scipy.io import loadmat
from settings import Settings
import matplotlib.pyplot as plt
from openpyxl.styles import Font
import matplotlib.gridspec as gridspec
from sklearn.metrics import confusion_matrix
from matplotlib.collections import LineCollection
from sklearn.model_selection import train_test_split
from tools import set_random_seed, load_dataset, standardize_eeg, get_path


def plot_eeg_shap_heatmap(
        e:np.ndarray, s:np.ndarray, chn:list, sr:int, p:dict, cmap:mpl.colors.Colormap,
        lt:int|None=None, lc:np.ndarray|None=None, acc:float|None=None
):
    """
    绘制EEG-SHAP热力图
    :param e: EEG数据, 形状(Channel, Time)
    :param s: SHAP数据, 形状(Channel, Time, Category)
    :param chn: 通道名称列表
    :param sr: 采样率
    :param p: 绘图参数
    :param cmap: 颜色图
    :param lt: 真实标签
    :param lc: 预测标签概率([为减速/0的概率, 为加速/1的概率])
    :param acc: 准确率
    """
    # region 准备数据
    ch_num, w_len = e.shape                                  # 通道数量与窗长
    cmap_range = float(max(abs(np.min(s)), abs(np.max(s))))  # 颜色图范围
    cmap_bin = np.linspace(-cmap_range, cmap_range, 256)
    # endregion
    # region 创建图窗
    fig = plt.figure(figsize=p['figure_size'], dpi=p['dpi'])
    gs = gridspec.GridSpec(1, 3, width_ratios=[1, 1, 0.05], figure=fig)
    ax0 = fig.add_subplot(gs[0, 0])
    ax1 = fig.add_subplot(gs[0, 1])
    cax = fig.add_subplot(gs[0, 2])
    # endregion
    # region 绘制子图
    x = np.arange(w_len)
    offsets = np.array(np.arange(ch_num-1, -1, -1))  # 用于分开每个通道, 免得挤在一起
    for subplot_idx in [0, 1]:  # 0:减速, 1: 加速
        if subplot_idx == 0: ax = ax0
        else: ax = ax1
        for ch_idx in range(ch_num):
            y = e[ch_idx, :] + offsets[ch_idx]
            segments, colors = [], []
            for t in range(w_len - 1):
                seg = [[x[t], y[t]], [x[t + 1], y[t + 1]]]  # 用左端点的 SHAP 值决定线段颜色
                col_idx = int(np.digitize(s[ch_idx, t, subplot_idx], cmap_bin)) - 1
                col_idx = np.clip(col_idx, 0, 255)
                colors.append(cmap(col_idx))
                segments.append(seg)
            linc = LineCollection(segments, colors=colors, linewidth=p['linewidth'], capstyle='round', joinstyle='round', antialiased=True)
            ax.add_collection(linc)

        ax.set_xlim(0, w_len - 1)  # 设置X轴
        ax.set_xticks([0, w_len - 1])
        ax.set_xticklabels(['0', f'{(w_len - 1)/sr}'], fontsize=p['xtick_fontsize'])
        ax.set_xlabel('Time (s)', fontsize=p['axis_label_fontsize'])

        ax.set_ylim(-1.5, offsets[0] + 1.5)
        ax.set_yticks(offsets)
        ax.set_yticklabels(chn, fontsize=p['ytick_fontsize'])

        if lc is not None:
            ax.set_title(f'Probability of {subplot_idx}: {lc[subplot_idx]:.5f}', fontsize=p['subtitle_fontsize'])
        else:
            ax.set_title(f'Label {subplot_idx}', fontsize=p['subtitle_fontsize'])

    # endregion
    # region 设置colorbar
    norm = mpl.colors.Normalize(vmin=-cmap_range, vmax=cmap_range)
    sm = mpl.cm.ScalarMappable(norm=norm, cmap=cmap)
    sm.set_array([])  # 虚拟数组，不需要实际数据
    cbar = fig.colorbar(sm, cax=cax, orientation='vertical')
    cbar.ax.set_title('SHAP value', fontsize=p['axis_label_fontsize'])
    cbar.ax.tick_params(labelsize=p['xtick_fontsize'])  # 可选，设置刻度字号
    # endregion
    if lt is not None and lc is not None:
        sup_title = f'Truth: {lt}, Prediction: {np.argmax(lc)}'  # 总标题
    else:
        sup_title = f'Accuracy: {acc:.5f}'
    fig.suptitle(sup_title, fontsize=p['main_title_fontsize'])
    plt.subplots_adjust(left=p['left'], right=p['right'], bottom=p['bottom'], top=p['top'], wspace=p['wspace'])
    return fig


def plot_shap_topomap(
        s:np.ndarray, chn:list, info:mne.Info, p:dict, cmap:mpl.colors.Colormap,
        lt:int|None=None, lc:np.ndarray|None=None, acc:float|None=None
):
    """
    绘制SHAP地形图
    :param s: 时间平均SHAP数据, 形状(Channel, Category)
    :param chn: 通道名称列表
    :param info: EEG信息, 用于通道定位
    :param p: 绘图参数
    :param cmap: 颜色图
    :param lt: 真实标签
    :param lc: 预测标签概率([为减速/0的概率, 为加速/1的概率])
    :param acc: 准确率
    """
    # region 准备数据
    cmap_range = float(max(abs(np.min(s)), abs(np.max(s))))  # 颜色图范围
    # endregion
    # region 创建图窗
    fig = plt.figure(figsize=p['figure_size'], dpi=p['dpi'])
    gs = gridspec.GridSpec(1, 3, width_ratios=[1, 1, 0.05], figure=fig)
    ax0 = fig.add_subplot(gs[0, 0])
    ax1 = fig.add_subplot(gs[0, 1])
    cax = fig.add_subplot(gs[0, 2])
    # endregion
    # region 绘制子图
    for subplot_idx in [0, 1]:  # 0:减速, 1: 加速
        if subplot_idx == 0: ax = ax0
        else: ax = ax1
        mne.viz.plot_topomap(
            data=s[:, subplot_idx],
            pos=info,
            names=chn,
            cmap=cmap,
            vlim=(-cmap_range, cmap_range),
            axes=ax,
            show=False
        )
        if lc is not None:
            ax.set_title(f'Probability of {subplot_idx}: {lc[subplot_idx]:.5f}', fontsize=p['subtitle_fontsize'])
        else:
            ax.set_title(f'Label {subplot_idx}', fontsize=p['subtitle_fontsize'])
    # endregion
    # region 设置colorbar
    norm = mpl.colors.Normalize(vmin=-cmap_range, vmax=cmap_range)
    sm = mpl.cm.ScalarMappable(norm=norm, cmap=cmap)
    sm.set_array([])                                    # 虚拟数组，不需要实际数据
    cbar = fig.colorbar(sm, cax=cax, orientation='vertical')
    cbar.ax.set_title('SHAP value', fontsize=p['axis_label_fontsize'])
    cbar.ax.tick_params(labelsize=p['xtick_fontsize'])  # 可选，设置刻度字号
    # endregion
    if lt is not None and lc is not None:
        sup_title = f'Truth: {lt}, Prediction: {np.argmax(lc)}'  # 总标题
    else:
        sup_title = f'Accuracy: {acc:.5f}'
    fig.suptitle(sup_title, fontsize=p['main_title_fontsize'])
    plt.subplots_adjust(left=p['left'], right=p['right'], bottom=p['bottom'], top=p['top'], wspace=p['wspace'])
    return fig


if __name__ == '__main__':
    # region 初始化
    set_random_seed()                                        # 设置随机种子
    S = Settings()                                           # 获取参数
    bandpass_filter_range = json.loads(sys.argv[1])          # 带通滤波频段
    colormaps = loadmat(get_path(path_spec='colormaps'))     # 载入自定义颜色图(包含多个颜色图)
    print(f'--------------------'
          f'Result analysis: '
          f'[{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]'
          f'--------------------')
    # endregion

    # region 模型效果
    if S.analysis_task['model_performance']:
        # region 测试集准确率表格
        # region 统计实验次数
        all_turns = set()  # 表格会分页, 每一页代表一次实验, 每一页上只会记录做了该次实验的受试
        for turns in S.subject_dictionary.values():
            all_turns.update(turns)
        all_turns = sorted(all_turns)
        # endregion
        # region 独立数据集准确率
        for network_name in S.network_list:
            print(f'Calculating test set accuracy of {network_name} [{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]...')
            sheets = {}
            for turn in all_turns:
                row_title = ([f'Imagine ({dtr[0]}-{dtr[1]})' for dtr in S.imagine_data_time_range_list] +
                             [f'Observe ({dtr[0]}-{dtr[1]})' for dtr in S.observe_data_time_range_list])  # 行标题
                subject_name_list = [
                    subject_name for subject_name, subject_turns in S.subject_dictionary.items()
                    if turn in subject_turns
                ]  # 进行过该次实验的受试(列标题)
                df = pd.DataFrame(index=row_title, columns=subject_name_list, dtype=float)

                for subject_name in subject_name_list:
                    for is_observe in [False, True]:
                        if is_observe:
                            data_time_range_list = S.observe_data_time_range_list
                        else:
                            data_time_range_list = S.imagine_data_time_range_list
                        for data_time_range in data_time_range_list:
                            # region 载入结果
                            path_truth = get_path(
                                path_spec='test_truth',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_truth):
                                print(f'    Missing file : {path_truth}')
                                continue
                            path_prediction = get_path(
                                path_spec='test_pred',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_prediction):
                                print(f'    Missing file : {path_prediction}')
                                continue
                            truth = np.load(path_truth)
                            prediction = np.load(path_prediction)
                            # endregion
                            # region 计算准确率并写入指定位置
                            row_idx = row_title.index(f'{"Observe" if is_observe else "Imagine"} ({data_time_range[0]}-{data_time_range[1]})')
                            column_idx = subject_name_list.index(subject_name)
                            accuracy = np.mean(np.argmax(prediction, axis=1) == truth)
                            df.iloc[row_idx, column_idx] = accuracy
                            # endregion
                if df.isnull().all().all(): continue
                sheets[str(turn)] = df
            # region 保存表格
            if not sheets:
                print(f'    No valid data for {network_name} [{bandpass_filter_range[0]}-{bandpass_filter_range[1]}].')
                continue
            path_accuracy_table = get_path(
                path_spec='acc_table',
                is_file=True,
                fr=bandpass_filter_range,
                net=network_name
            )
            with pd.ExcelWriter(path_accuracy_table, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name)
                    ws = writer.book[sheet_name]
                    ws.column_dimensions['A'].width = S.accuracy_table_param['first_column_width']
                    # region 根据准确率为文本着色
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                        for cell in row:
                            val = cell.value
                            if isinstance(val, (int, float)):
                                if val >= S.accuracy_table_param['good_accuracy_threshold']:
                                    cell.font = Font(color=S.accuracy_table_param['good_accuracy_color'])
                                elif val <= S.accuracy_table_param['bad_accuracy_threshold']:
                                    cell.font = Font(color=S.accuracy_table_param['bad_accuracy_color'])
                    # endregion
            print(f'    Accuracy table has been saved to: {path_accuracy_table}')
            # endregion
        # endregion
        # region 合并数据集准确率
        for network_name in S.network_list:
            print(f'Calculating combined test set accuracy of {network_name} [{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]...')
            sheets = {}
            for turn in all_turns:
                row_title = ([f'Imagine ({dtr[0]}-{dtr[1]})' for dtr in S.imagine_data_time_range_list] +
                             [f'Observe ({dtr[0]}-{dtr[1]})' for dtr in S.observe_data_time_range_list])  # 行标题
                subject_name_list = [
                    subject_name for subject_name, subject_turns in S.subject_dictionary.items()
                    if turn in subject_turns
                ]  # 进行过该次实验的受试(列标题)
                df = pd.DataFrame(index=row_title, columns=subject_name_list, dtype=float)

                turn_list = [t for t in all_turns if t <= turn]
                if len(turn_list) == 1: turn_list = turn_list[0]

                for subject_name in subject_name_list:
                    for is_observe in [False, True]:
                        if is_observe:
                            data_time_range_list = S.observe_data_time_range_list
                        else:
                            data_time_range_list = S.imagine_data_time_range_list
                        for data_time_range in data_time_range_list:
                            # region 载入结果
                            path_truth = get_path(
                                path_spec='test_truth',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=turn_list,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_truth):
                                print(f'    Missing file : {path_truth}')
                                continue
                            path_prediction = get_path(
                                path_spec='test_pred',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=turn_list,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_prediction):
                                print(f'    Missing file : {path_prediction}')
                                continue
                            truth = np.load(path_truth)
                            prediction = np.load(path_prediction)
                            # endregion
                            # region 计算准确率并写入指定位置
                            row_idx = row_title.index(f'{"Observe" if is_observe else "Imagine"} ({data_time_range[0]}-{data_time_range[1]})')
                            column_idx = subject_name_list.index(subject_name)
                            accuracy = np.mean(np.argmax(prediction, axis=1) == truth)
                            df.iloc[row_idx, column_idx] = accuracy
                            # endregion
                if df.isnull().all().all(): continue
                sheets[str(turn_list)] = df
            # region 保存表格
            if not sheets:
                print(f'    No valid data for {network_name} [{bandpass_filter_range[0]}-{bandpass_filter_range[1]}].')
                continue
            path_accuracy_table = get_path(
                path_spec='comb_acc_table',
                is_file=True,
                fr=bandpass_filter_range,
                net=network_name
            )
            with pd.ExcelWriter(path_accuracy_table, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name.replace('[', '').replace(']', ''))
                    ws = writer.book[sheet_name.replace('[', '').replace(']', '')]
                    ws.column_dimensions['A'].width = S.accuracy_table_param['first_column_width']
                    # region 根据准确率为文本着色
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                        for cell in row:
                            val = cell.value
                            if isinstance(val, (int, float)):
                                if val >= S.accuracy_table_param['good_accuracy_threshold']:
                                    cell.font = Font(color=S.accuracy_table_param['good_accuracy_color'])
                                elif val <= S.accuracy_table_param['bad_accuracy_threshold']:
                                    cell.font = Font(color=S.accuracy_table_param['bad_accuracy_color'])
                    # endregion
            print(f'    Combined accuracy table has been saved to: {path_accuracy_table}')
            # endregion
        # endregion
        # endregion
        # region 测试集混淆矩阵
        # region 独立数据集混淆矩阵
        print(f'Calculating test set confusion matrix of [{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]...')
        for subject_name in S.subject_dictionary:
            for subject_turn in S.subject_dictionary[subject_name]:
                for is_observe in S.use_observe_data:
                    if is_observe: data_time_range_list = S.observe_data_time_range_list
                    else: data_time_range_list = S.imagine_data_time_range_list
                    for data_time_range in data_time_range_list:
                        for network_name in S.network_list:
                            # region 载入真实与预测标签
                            path_truth = get_path(
                                path_spec='test_truth',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_truth):
                                print(f'    Missing file : {path_truth}')
                                continue
                            path_prediction = get_path(
                                path_spec='test_pred',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_prediction):
                                print(f'    Missing file : {path_prediction}')
                                continue
                            truth = np.load(path_truth)
                            prediction = np.argmax(np.load(path_prediction), axis=1)
                            # endregion
                            # region 计算, 绘制, 保存混淆矩阵
                            matrix = confusion_matrix(truth, prediction)
                            df = pd.DataFrame(matrix, index=['True 0', 'True 1'], columns=['Pred 0', 'Pred 1'])
                            path_confusion_matrix = get_path(
                                path_spec='confusion_matrix',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=data_time_range,
                                net=network_name
                            )
                            df.to_excel(path_confusion_matrix)
                            print(f'    Confusion matrix of {subject_name}, '
                                  f'{subject_turn}, '
                                  f'{"Observe" if is_observe else "Imagine"}, '
                                  f'({data_time_range[0]}-{data_time_range[1]}), '
                                  f'{network_name} has been saved to : {path_confusion_matrix}')
                            # endregion
        # endregion
        # region TODO 1: 合并数据集混淆矩阵
        print(f'Calculating combined test set confusion matrix of [{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]...')

        # endregion
        # endregion
    # endregion

    # region TODO 2: 神经表征
    if S.analysis_task['neural_representation']:
        if bandpass_filter_range != S.neural_representation_target_band:
            print(f'Skipped neural representation for non-target band '
                  f'[{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]Hz.')
        else:
            print(f'Visualizing neural representation for target band '
                  f'[{bandpass_filter_range[0]}-{bandpass_filter_range[1]}]Hz...')
            for subject_name in S.subject_dictionary:
                for subject_turn in S.subject_dictionary[subject_name]:
                    for is_observe in [False, True]:
                        # region 载入预处理数据
                        path_neural_data = get_path(
                            path_spec='neu_rep_data',
                            s_name=subject_name,
                            s_turn=subject_turn,
                            obs=is_observe,
                            fr=bandpass_filter_range,
                        )
                        try:
                            num_block = len([
                                f for f in os.listdir(path_neural_data)
                                if os.path.isfile(os.path.join(path_neural_data, f))
                            ])
                        except FileNotFoundError:
                            print(f'    Missing data : {path_neural_data}')
                            continue

                        print(f'    {subject_name}-{subject_turn}-{"Observe" if is_observe else "Imagine"} has {num_block} blocks.')
                        block_list = []
                        for block_idx in range(num_block):
                            block = mne.io.read_raw_fif(
                                os.path.join(path_neural_data, f'B{int(block_idx + 1)}_raw.fif'),
                                preload=True,
                                verbose=0
                            )
                            block_list.append(block)
                        # endregion
                        # region 提取试次
                        total_num_obs_acc_trig = 0  # 观察加速标签总数
                        total_num_obs_dec_trig = 0  # 观察减速标签总数
                        total_num_imag_trig = 0  # 想象标签总数
                        total_num_epoch = 0  # 试次总数
                        total_num_bad_trial = 0  # 坏段总数
                        epoched_block_list = []  # 记录所有提取过试次的组块(已剔除伪迹)

                        for block_idx, block in enumerate(block_list):
                            block_events, event_dict = mne.events_from_annotations(block)        # 获取标签列表
                            obs_acc_trig_id = event_dict.get(str(S.trigger_observe_accelerate))  # 观察加速标签ID
                            obs_dec_trig_id = event_dict.get(str(S.trigger_observe_decelerate))  # 观察减速标签ID
                            imag_trig_id = event_dict.get(str(S.trigger_imagine))                # 想象标签ID

                            target_events = []  # 记录所有想象标签
                            for event_idx, event in enumerate(block_events):
                                if event[2] == imag_trig_id:
                                    target_events.append(event)
                            target_events = np.array(target_events, dtype=int)

                            epoched_block = mne.Epochs(  # 提取完整试次
                                block, target_events, event_id=imag_trig_id,
                                tmin=-(S.time_prepare + S.time_observe), tmax=S.time_imagine,
                                baseline=(-(S.time_prepare + S.time_observe), -S.time_observe) if is_observe else (-S.baseline_time, 0),  # 基线范围, 注意这里不会删除基线时间段
                                picks='eeg', preload=True
                            )
                        # endregion
    # endregion

    # region 特征贡献
    if S.analysis_task['feature_contribution']:
        print(f'Plotting SHAP features...')
        # region 设置颜色图
        if S.eeg_shap_heatmap_param['colormap'] not in colormaps.keys():
            print(f'    Colormap "{S.eeg_shap_heatmap_param["colormap"]}" not found in '
                  f'{get_path(path_spec="colormaps")}, subprocess ended.')
            exit()
        eeg_shap_colormap = mpl.colors.ListedColormap(colormaps[S.eeg_shap_heatmap_param['colormap']], name=S.eeg_shap_heatmap_param['colormap'])
        if S.shap_topomap_param['colormap'] not in colormaps.keys():
            print(f'    Colormap "{S.shap_topomap_param["colormap"]}" not found in '
                  f'{get_path(path_spec="colormaps")}, subprocess ended.')
            exit()
        shap_topo_colormap = mpl.colors.ListedColormap(colormaps[S.shap_topomap_param['colormap']], name=S.shap_topomap_param['colormap'])
        # endregion
        for subject_name in S.subject_dictionary:
            skip_turns = False  # 如果S.concat_turns为真, 则只会循环一次subject_turn
            for subject_turn in S.subject_dictionary[subject_name]:
                if skip_turns:
                    continue
                if S.concat_turns and len(S.subject_dictionary[subject_name]) > 1:
                    subject_turn = S.subject_dictionary[subject_name]
                # region 设置通道位置
                if isinstance(subject_turn, int):
                    path_chan_loc = get_path(path_spec='chan_loc', s_name=subject_name, s_turn=subject_turn)
                else:
                    path_chan_loc = get_path(path_spec='chan_loc', s_name=subject_name, s_turn=subject_turn[0])  # TODO 5: 换用其他帽子后会不兼容吗
                if not os.path.exists(path_chan_loc):
                    print (f'    Missing channel location file : {path_chan_loc}')
                    continue
                montage = mne.channels.read_custom_montage(path_chan_loc)  # 读取每个受试每次实验的通道位置(此时不包含基准点)

                is_eeg_channel = [ch not in S.non_eeg_channel for ch in montage.ch_names]                  # 通道类型是否为EEG
                montage.ch_names = [ch for ch, is_eeg in zip(montage.ch_names, is_eeg_channel) if is_eeg]  # 只保留EEG通道
                montage.dig = [d for d, is_eeg in zip(montage.dig, is_eeg_channel) if is_eeg]

                for channel_idx in range(len(montage.ch_names)):  # 将每个通道投影至一个球体上, 便于地形图的绘制
                    coord = montage.dig[channel_idx]['r']
                    r = float(np.linalg.norm(coord))
                    if r > 0:
                        montage.dig[channel_idx]['r'] = (coord / r) * S.shap_topomap_sphere_radius

                cz_loc = montage.dig[montage.ch_names.index('Cz')]['r']  # 获取Cz位置
                for channel_idx in range(len(montage.ch_names)):              # 将Cz移至轮廓中央
                    montage.dig[channel_idx]['r'] = montage.dig[channel_idx]['r'] - cz_loc

                montage.dig += mne.channels.make_standard_montage(S.montage).dig[0:3]  # 从标准电极位置中获取基准点(LPA, Nasion, RPA)
                mne.channels.make_dig_montage()
                eeg_info = mne.create_info(ch_names=montage.ch_names, sfreq=S.new_sampling_rate, ch_types='eeg')
                eeg_info.set_montage(montage)
                # endregion
                for is_observe in S.use_observe_data:
                    if is_observe:
                        data_time_range_list = S.observe_data_time_range_list
                    else:
                        data_time_range_list = S.imagine_data_time_range_list
                    for window_time_range in data_time_range_list:
                        for network_name in S.network_list:
                            # region 载入EEG与标签
                            try:
                                eeg, label, bad_trial = load_dataset(subject_name, subject_turn, is_observe, bandpass_filter_range)  # TODO 4: 兼容合并数据集
                            except FileNotFoundError:
                                print(f'    Missing data : {subject_name}, '
                                      f'{subject_turn}, '
                                      f'{"observe" if is_observe else "imagine"}, '
                                      f'[{bandpass_filter_range[0]}-{bandpass_filter_range[1]}].')
                                continue
                            if eeg.shape[1] != len(montage.ch_names):
                                print (f'EEG data incompatible with location file : {path_chan_loc}')
                                continue
                            bad_trial_indices = [
                                int((bad_trial[idx, 0] - 1) * S.num_trial_subject_turn[subject_name][str(subject_turn)] + (bad_trial[idx, 1] - 1))
                                for idx in range(bad_trial.shape[0])
                            ]                        # 计算坏段索引
                            num_deleted_trial = len(bad_trial_indices)       # 坏段数量
                            eeg = np.delete(eeg, bad_trial_indices, axis=0)  # 删除坏段
                            label = np.delete(label, bad_trial_indices, axis=0)
                            eeg = eeg[:, :,
                                      int(S.new_sampling_rate * window_time_range[0]):int(S.new_sampling_rate * window_time_range[1])]                                   # 截取所需时间段
                            path_test_truth = get_path(                      # 真实标签路径
                                path_spec='test_truth',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=window_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_test_truth):
                                print(f'    Missing file : {path_test_truth}')
                                continue
                            test_truth = np.load(path_test_truth)            # 载入真实标签
                            path_test_prediction = get_path(                 # 预测标签路径
                                path_spec='test_pred',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=window_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_test_prediction):
                                print(f'    Missing file : {path_test_prediction}')
                                continue
                            test_prediction = np.load(path_test_prediction)  # 载入预测标签
                            # endregion
                            # region 划分数据集
                            training_eeg, test_eeg, training_label, _ = train_test_split(  # 划分出测试集用于绘图(只要随机种子相同, 这里的测试集EEG就会与上面载入的标签相对应)
                                eeg, label, test_size=S.test_ratio, shuffle=True, random_state=S.random_seed
                            )
                            train_eeg, validation_eeg, _, _ = train_test_split(                     # 划分出训练集与验证集用于标准化
                                training_eeg, training_label, test_size=S.test_ratio, shuffle=True,
                                random_state=S.random_seed
                            )
                            _, _, test_eeg = standardize_eeg(train_eeg, validation_eeg, test_eeg)   # 标准化测试集, 形状(Trial, Channel, Time), (Trial,)
                            del eeg, label, bad_trial, training_eeg, training_label, validation_eeg
                            # endregion
                            # region 载入SHAP值
                            path_explanation = get_path(             # SHAP值路径
                                path_spec='explanation',
                                is_file=True,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=window_time_range,
                                net=network_name
                            )
                            if not os.path.exists(path_explanation):
                                print(f'    Missing file : {path_explanation}')
                                continue
                            explanation = np.load(path_explanation)  # 形状(Trial, Channel, Time, Category)
                            if explanation.shape[:3] != test_eeg.shape:
                                print(f'    {subject_name}, '
                                      f'{subject_turn}, '
                                      f'{"Observe" if is_observe else "Imagine"}, '
                                      f'[{bandpass_filter_range[0]}-{bandpass_filter_range[1]}], '
                                      f'({window_time_range[0]}-{window_time_range[1]}), '
                                      f'{network_name} : '
                                      f'EEG data {test_eeg.shape} and SHAP data {explanation.shape} are incompatible in shape, skipped.')
                                continue
                            num_trial = explanation.shape[0]
                            # endregion
                            # region 创建保存路径
                            path_eeg_shap_heatmap = get_path(
                                path_spec='eeg_shap_heatmap',
                                is_file=False,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=window_time_range,
                                net=network_name
                            )
                            path_shap_topomap = get_path(
                                path_spec='shap_topomap',
                                is_file=False,
                                s_name=subject_name,
                                s_turn=subject_turn,
                                obs=is_observe,
                                fr=bandpass_filter_range,
                                tr=window_time_range,
                                net=network_name
                            )
                            os.makedirs(path_eeg_shap_heatmap, exist_ok=True)
                            os.makedirs(path_shap_topomap, exist_ok=True)
                            # endregion
                            # region 绘制EEG-SHAP图
                            num_skipped_plot = 0
                            for trial_idx in range(num_trial):
                                path_plot = os.path.join(path_eeg_shap_heatmap, f'Trial {trial_idx + 1}.png')
                                if not os.path.exists(path_plot) or S.eeg_shap_heatmap_param['overwrite']:
                                    figure = plot_eeg_shap_heatmap(
                                        e=test_eeg[trial_idx],
                                        s=explanation[trial_idx],
                                        chn=montage.ch_names,
                                        sr=S.new_sampling_rate,
                                        p=S.eeg_shap_heatmap_param,
                                        cmap=eeg_shap_colormap,
                                        lt=int(test_truth[trial_idx]),
                                        lc=test_prediction[trial_idx],
                                    )
                                    figure.savefig(path_plot)
                                    plt.close(figure)
                                else:
                                    num_skipped_plot += 1
                            path_plot = os.path.join(path_eeg_shap_heatmap, f'Trial Average.png')
                            if not os.path.exists(path_plot) or S.eeg_shap_heatmap_param['overwrite']:
                                average_plot_params = S.eeg_shap_heatmap_param.copy()
                                average_plot_params['linewidth'] = average_plot_params['average_linewidth']
                                figure = plot_eeg_shap_heatmap(
                                    e=np.zeros_like(test_eeg[0]),
                                    s=np.mean(explanation, axis=0),
                                    chn=montage.ch_names,
                                    sr=S.new_sampling_rate,
                                    p=average_plot_params,
                                    cmap=eeg_shap_colormap,
                                    acc=float(np.mean(np.argmax(test_prediction, axis=1) == test_truth))
                                )
                                figure.savefig(path_plot)
                                plt.close(figure)
                            else:
                                num_skipped_plot += 1

                            hint_text = (f'    {subject_name}, '
                                         f'{subject_turn}, '
                                         f'{"Observe" if is_observe else "Imagine"}, '
                                         f'[{bandpass_filter_range[0]}-{bandpass_filter_range[1]}], '
                                         f'({window_time_range[0]}-{window_time_range[1]}), '
                                         f'{network_name} : ')
                            if 0 < num_skipped_plot:
                                print(f'{hint_text}'
                                      f'Skipped {num_skipped_plot} figure{"" if num_skipped_plot == 1 else "s"} '
                                      f'because {"it" if num_skipped_plot == 1 else "they"} exist{"" if num_skipped_plot == 1 else "s"}.')
                            elif S.eeg_shap_heatmap_param['overwrite']:
                                print(f'{hint_text}Overwrote all existing figures.')
                            if num_skipped_plot < num_trial:
                                print(f'{hint_text}EEG-SHAP heatmaps have been saved to {path_eeg_shap_heatmap}')
                            # endregion
                            # region 绘制SHAP地形图
                            num_skipped_plot = 0
                            for trial_idx in range(num_trial):
                                path_plot = os.path.join(path_shap_topomap, f'Trial {trial_idx + 1}.png')
                                if not os.path.exists(path_plot) or S.shap_topomap_param['overwrite']:
                                    figure = plot_shap_topomap(
                                        s=np.mean(explanation[trial_idx], axis=1),
                                        chn=montage.ch_names,
                                        info=eeg_info,
                                        p=S.shap_topomap_param,
                                        cmap=shap_topo_colormap,
                                        lt=int(test_truth[trial_idx]),
                                        lc=test_prediction[trial_idx]
                                    )
                                    figure.savefig(path_plot)
                                    plt.close(figure)
                                else:
                                    num_skipped_plot += 1
                            path_plot = os.path.join(path_shap_topomap, f'Trial Average.png')
                            if not os.path.exists(path_plot) or S.shap_topomap_param['overwrite']:
                                figure = plot_shap_topomap(
                                    s=np.mean(np.mean(explanation, axis=2), axis=0),
                                    chn=montage.ch_names,
                                    info=eeg_info,
                                    p=S.shap_topomap_param,
                                    cmap=shap_topo_colormap,
                                    acc=float(np.mean(np.argmax(test_prediction, axis=1) == test_truth))
                                )
                                figure.savefig(path_plot)
                                plt.close(figure)
                            else:
                                num_skipped_plot += 1

                            if num_skipped_plot > 0:
                                print(f'{hint_text}'
                                      f'Skipped {num_skipped_plot} figure{"" if num_skipped_plot == 1 else "s"} '
                                      f'because {"it" if num_skipped_plot == 1 else "they"} exist{"" if num_skipped_plot == 1 else "s"}.')
                            elif S.eeg_shap_heatmap_param['overwrite']:
                                print(f'{hint_text}Overwrote all existing figures.')
                            if num_skipped_plot < num_trial:
                                print(f'{hint_text}SHAP topography maps have been saved to {path_shap_topomap}')
                            # endregion

                if S.concat_turns:
                    skip_turns = True
    # endregion
